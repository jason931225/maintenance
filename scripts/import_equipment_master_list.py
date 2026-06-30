#!/usr/bin/env python3
"""Generate safe SQL for the flat Korean equipment master workbook.

The production registry importer accepts the older two-sheet K&L workbook.  The
current live source from TalkFile is a flat `Sheet2` export, so this generator
maps that file into the same governed registry tables while preserving the
source-only columns in `registry_equipment.note` as compact JSON.

PII/data safety:
  * stdout contains only aggregate counts and output paths.
  * generated SQL contains operational/source data; write it under `.omx/` with
    mode 0600 and never commit it.
  * dry-run SQL wraps all mutations in ROLLBACK; apply SQL commits only after the
    same validations pass inside the transaction.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import numbers
import os
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEFAULT_WORKBOOK = Path(
    "/Users/jasonlee/Library/Mobile Documents/com~apple~CloudDocs/"
    "TalkFile_장비Master List.xlsx"
)
KNL_ORG_ID = "00000000-0000-0000-0000-0000000000a1"

REQUIRED_HEADERS = [
    "장비 No",
    "사업장",
    "계약처",
    "상태",
    "담당자",
    "배치장소",
    "배치No",
    "운영시간",
    "규격",
    "인양능력(톤)",
    "인양높이",
    "제작처",
    "모델명",
    "차대번호",
    "년식",
    "가동시간",
    "차량 No.",
    "보험",
    "보험사",
    "계약자",
    "피보험자",
    "자산처",
    "자산등록일",
    "임대시작일",
    "임대료",
    "차량가액",
    "상각년수",
    "잔존가치",
    "차량수리비",
    "임대처",
]

STATUS = {
    "임대": "임대",
    "예비": "예비",
    "폐기": "폐기",
    "대체": "대체",
    "매각": "매각",
}
EQUIPMENT_NO_RE = re.compile(r"^[A-Z]{3}[A-Z0-9]{2}-[0-9]{4}$")


def nfc(value: str) -> str:
    return unicodedata.normalize("NFC", value)


def cell_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, bool):
        return value
    if isinstance(value, numbers.Integral):
        return int(value)
    if isinstance(value, numbers.Real):
        return int(value) if float(value).is_integer() else float(value)
    return nfc(str(value)).strip()


def text(value: Any) -> str:
    value = cell_value(value)
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def opt_text(value: Any) -> str | None:
    t = text(value)
    return t if t else None


def parse_date(value: Any, *, field: str, row: int) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, numbers.Real):
        # Excel serial date. 1899-12-30 matches openpyxl's serial convention for
        # modern workbooks; this branch is a fallback because openpyxl usually
        # gives datetime/date for date-formatted cells.
        return (dt.date(1899, 12, 30) + dt.timedelta(days=int(value))).isoformat()
    raw = text(value)
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d", "%Y"):
        try:
            parsed = dt.datetime.strptime(raw, fmt).date()
            return parsed.isoformat()
        except ValueError:
            pass
    raise ValueError(f"row {row}: invalid {field} date {raw!r}")


def parse_int(value: Any, *, field: str, row: int) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, numbers.Integral):
        return int(value)
    if isinstance(value, numbers.Real):
        return int(round(float(value)))
    raw = re.sub(r"[,_\s원]", "", text(value))
    if not raw:
        return None
    try:
        return int(float(raw))
    except ValueError as exc:
        raise ValueError(f"row {row}: invalid {field} integer {text(value)!r}") from exc


def parse_bool_yn(value: Any, *, field: str, row: int) -> bool | None:
    raw = text(value).upper()
    if not raw:
        return None
    if raw in {"Y", "YES", "TRUE", "1", "O", "○"}:
        return True
    if raw in {"N", "NO", "FALSE", "0", "X"}:
        return False
    raise ValueError(f"row {row}: invalid {field} Y/N value {raw!r}")


def sql_str(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + value.replace("'", "''") + "'"


def sql_uuid(value: str | None) -> str:
    return "NULL" if value is None else sql_str(value) + "::uuid"


def sql_date(value: str | None) -> str:
    return "NULL" if value is None else sql_str(value) + "::date"


def sql_int(value: int | None) -> str:
    return "NULL" if value is None else str(value)


def sql_bool(value: bool | None) -> str:
    if value is None:
        return "NULL"
    return "TRUE" if value else "FALSE"


def sql_json(value: Any) -> str:
    return sql_str(json.dumps(value, ensure_ascii=False, separators=(",", ":"))) + "::jsonb"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def ton_milli(ton_text: str) -> int | None:
    raw = ton_text.strip().upper()
    if not raw.endswith("T"):
        return None
    try:
        return int(round(float(raw[:-1]) * 1000))
    except ValueError:
        return None


def equipment_codes(equipment_no: str) -> tuple[str, str, str]:
    return equipment_no[0:1], equipment_no[1:2], equipment_no[2:3]


def source_note(row: dict[str, Any], workbook_name: str, source_row: int) -> str:
    payload = {
        "source_file": workbook_name,
        "source_sheet": "Sheet2",
        "source_row": source_row,
        "operating_org": "knl",
        "source_columns": {
            "사업장": opt_text(row.get("사업장")),
            "계약처": opt_text(row.get("계약처")),
            "인양높이": opt_text(row.get("인양높이")),
            "상각년수": cell_value(row.get("상각년수")),
            "차량수리비": cell_value(row.get("차량수리비")),
            "임대처": opt_text(row.get("임대처")),
        },
    }
    return json.dumps(payload, ensure_ascii=False, separators=(",", ":"))


def load_rows(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Sheet2" not in wb.sheetnames:
        raise ValueError(f"expected Sheet2 in workbook, found {wb.sheetnames!r}")
    ws = wb["Sheet2"]
    header_values = [text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if header_values[: len(REQUIRED_HEADERS)] != REQUIRED_HEADERS:
        raise ValueError(
            "unexpected equipment workbook headers; refusing lossy import: "
            + json.dumps(header_values, ensure_ascii=False)
        )

    rows: list[dict[str, Any]] = []
    seen: set[str] = set()
    for source_row, cells in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        raw = dict(zip(REQUIRED_HEADERS, cells))
        if not any(text(value) for value in raw.values()):
            continue
        equipment_no = text(raw.get("장비 No"))
        if not EQUIPMENT_NO_RE.match(equipment_no):
            raise ValueError(f"row {source_row}: invalid 장비 No {equipment_no!r}")
        if equipment_no in seen:
            raise ValueError(f"row {source_row}: duplicate 장비 No {equipment_no!r}")
        seen.add(equipment_no)
        status = text(raw.get("상태"))
        if status not in STATUS:
            raise ValueError(f"row {source_row}: unknown 상태 {status!r}")
        site_name = opt_text(raw.get("사업장")) or opt_text(raw.get("배치장소"))
        customer_name = opt_text(raw.get("계약처")) or site_name
        specification = opt_text(raw.get("규격"))
        ton_text = opt_text(raw.get("인양능력(톤)"))
        if not site_name or not customer_name or not specification or not ton_text:
            raise ValueError(f"row {source_row}: missing required customer/site/spec/ton")
        manufacturer_code, kind_code, power_code = equipment_codes(equipment_no)
        rows.append(
            {
                "source_row": source_row,
                "equipment_no": equipment_no,
                "branch_id": None,  # filled by SQL target branch argument
                "customer_name": customer_name,
                "site_name": site_name,
                "management_no": opt_text(raw.get("배치No")),
                "manufacturer_code": manufacturer_code,
                "kind_code": kind_code,
                "power_code": power_code,
                "power_label": opt_text(raw.get("인양높이")),
                "status": status,
                "manager_name": opt_text(raw.get("담당자")),
                "placement_location": opt_text(raw.get("배치장소")),
                "placement_no": opt_text(raw.get("배치No")),
                "operation_shift": opt_text(raw.get("운영시간")),
                "specification": specification,
                "ton_text": ton_text,
                "ton_milli": ton_milli(ton_text),
                "maker": opt_text(raw.get("제작처")),
                "model": opt_text(raw.get("모델명")),
                "vin": opt_text(raw.get("차대번호")),
                "year": parse_date(raw.get("년식"), field="년식", row=source_row),
                "hours": parse_int(raw.get("가동시간"), field="가동시간", row=source_row),
                "vehicle_registration_no": opt_text(raw.get("차량 No.")),
                "insured": parse_bool_yn(raw.get("보험"), field="보험", row=source_row),
                "insurer": opt_text(raw.get("보험사")),
                "policy_holder": opt_text(raw.get("계약자")),
                "insured_party": opt_text(raw.get("피보험자")),
                "asset_owner": opt_text(raw.get("자산처")),
                "asset_registered_on": parse_date(raw.get("자산등록일"), field="자산등록일", row=source_row),
                "rental_started_on": parse_date(raw.get("임대시작일"), field="임대시작일", row=source_row),
                "rental_fee": parse_int(raw.get("임대료"), field="임대료", row=source_row),
                "vehicle_value": parse_int(raw.get("차량가액"), field="차량가액", row=source_row),
                "residual_value": parse_int(raw.get("잔존가치"), field="잔존가치", row=source_row),
                "note": source_note(raw, path.name, source_row),
                "source_sheet": "Sheet2",
            }
        )
    return rows


def upsert_sql(row: dict[str, Any], org_id: str, branch_id: str) -> str:
    equipment_columns = [
        "equipment_no",
        "branch_id",
        "customer_id",
        "site_id",
        "management_no",
        "manufacturer_code",
        "kind_code",
        "power_code",
        "power_label",
        "status",
        "manager_name",
        "placement_location",
        "placement_no",
        "operation_shift",
        "specification",
        "ton_text",
        "ton_milli",
        "maker",
        "model",
        "vin",
        "year",
        "hours",
        "vehicle_registration_no",
        "insured",
        "insurer",
        "policy_holder",
        "insured_party",
        "asset_owner",
        "asset_registered_on",
        "rental_started_on",
        "rental_fee",
        "vehicle_value",
        "residual_value",
        "note",
        "source_sheet",
        "source_row",
        "org_id",
    ]
    update_columns = [c for c in equipment_columns if c not in {"equipment_no", "org_id"}]
    select_values = [
        sql_str(row["equipment_no"]),
        "target.branch_id",
        "cust.id",
        "site.id",
        sql_str(row["management_no"]),
        sql_str(row["manufacturer_code"]),
        sql_str(row["kind_code"]),
        sql_str(row["power_code"]),
        sql_str(row["power_label"]),
        sql_str(row["status"]),
        sql_str(row["manager_name"]),
        sql_str(row["placement_location"]),
        sql_str(row["placement_no"]),
        sql_str(row["operation_shift"]),
        sql_str(row["specification"]),
        sql_str(row["ton_text"]),
        sql_int(row["ton_milli"]),
        sql_str(row["maker"]),
        sql_str(row["model"]),
        sql_str(row["vin"]),
        sql_date(row["year"]),
        sql_int(row["hours"]),
        sql_str(row["vehicle_registration_no"]),
        sql_bool(row["insured"]),
        sql_str(row["insurer"]),
        sql_str(row["policy_holder"]),
        sql_str(row["insured_party"]),
        sql_str(row["asset_owner"]),
        sql_date(row["asset_registered_on"]),
        sql_date(row["rental_started_on"]),
        sql_int(row["rental_fee"]),
        sql_int(row["vehicle_value"]),
        sql_int(row["residual_value"]),
        sql_str(row["note"]),
        sql_str(row["source_sheet"]),
        str(row["source_row"]),
        "target.org_id",
    ]
    return f"""
WITH target AS (
    SELECT {sql_uuid(org_id)} AS org_id, {sql_uuid(branch_id)} AS branch_id
), cust AS (
    INSERT INTO registry_customers (branch_id, name, org_id)
    SELECT target.branch_id, {sql_str(row['customer_name'])}, target.org_id FROM target
    ON CONFLICT (branch_id, name) DO UPDATE SET updated_at = registry_customers.updated_at
    RETURNING id
), site AS (
    INSERT INTO registry_sites (branch_id, customer_id, name, org_id)
    SELECT target.branch_id, cust.id, {sql_str(row['site_name'])}, target.org_id FROM target, cust
    ON CONFLICT (branch_id, customer_id, name) DO UPDATE SET updated_at = registry_sites.updated_at
    RETURNING id
)
INSERT INTO registry_equipment ({', '.join(equipment_columns)})
SELECT {', '.join(select_values)}
FROM target, cust, site
ON CONFLICT (org_id, equipment_no) DO UPDATE SET
    {', '.join(f'{col} = EXCLUDED.{col}' for col in update_columns)},
    updated_at = now();
""".strip()


def build_sql(rows: list[dict[str, Any]], *, workbook: Path, org_id: str, branch_id: str, mode: str) -> str:
    source_hash = sha256_file(workbook)
    status_counts = Counter(row["status"] for row in rows)
    owner_counts = Counter(row["asset_owner"] or "미지정" for row in rows)
    summary = {
        "source": workbook.name,
        "source_sha256": source_hash,
        "format": "flat_equipment_master_sheet2",
        "mode": mode,
        "target_org": "knl",
        "target_org_id": org_id,
        "target_branch_id": branch_id,
        "input_rows": len(rows),
        "equipment_count": len(rows),
        "status_counts": dict(sorted(status_counts.items())),
        "asset_owner_counts": dict(sorted(owner_counts.items())),
        "legal_owner_preserved_in": "registry_equipment.asset_owner",
        "operating_org_policy": "all rows imported under KNL org/branch so KNL EquipmentManage users can operate the group fleet while source legal owner remains preserved",
    }
    footer = "ROLLBACK;" if mode == "dry-run" else "COMMIT;"
    parts = [
        "\\set ON_ERROR_STOP on",
        "SET ROLE mnt_rt;",
        "BEGIN;",
        f"SELECT set_config('app.current_org', {sql_str(org_id)}, true);",
        "SELECT 'target_org', slug, id FROM organizations WHERE id = " + sql_uuid(org_id) + ";",
        "SELECT 'target_branch', name, id FROM branches WHERE id = " + sql_uuid(branch_id) + " AND org_id = " + sql_uuid(org_id) + ";",
        "DO $$ BEGIN IF NOT EXISTS (SELECT 1 FROM branches WHERE id = " + sql_uuid(branch_id) + " AND org_id = " + sql_uuid(org_id) + ") THEN RAISE EXCEPTION 'target KNL branch not found'; END IF; END $$;",
        "SELECT 'before_equipment_count', COUNT(*) FROM registry_equipment;",
        "SELECT 'before_target_branch_count', COUNT(*) FROM registry_equipment WHERE branch_id = " + sql_uuid(branch_id) + ";",
    ]
    parts.extend(upsert_sql(row, org_id, branch_id) for row in rows)
    parts.extend(
        [
            "INSERT INTO audit_events (actor, action, target_type, target_id, branch_id, before_snap, after_snap, trace_id, span_id, occurred_at, org_id) "
            "SELECT NULL, 'registry.import', 'registry_import', "
            + sql_str(workbook.name)
            + ", "
            + sql_uuid(branch_id)
            + ", NULL, "
            + sql_json(summary)
            + ", lower(replace(gen_random_uuid()::text, '-', '')), substring(lower(replace(gen_random_uuid()::text, '-', '')) from 1 for 16), now(), "
            + sql_uuid(org_id)
            + ";",
            "SELECT 'after_equipment_count', COUNT(*) FROM registry_equipment;",
            "SELECT 'after_imported_source_count', COUNT(*) FROM registry_equipment WHERE source_sheet = 'Sheet2' AND note LIKE '%TalkFile_%Master List.xlsx%';",
            "SELECT 'after_target_branch_count', COUNT(*) FROM registry_equipment WHERE branch_id = " + sql_uuid(branch_id) + ";",
            "SELECT 'after_legal_owner_values', COUNT(DISTINCT asset_owner) FROM registry_equipment WHERE source_sheet = 'Sheet2' AND note LIKE '%TalkFile_%Master List.xlsx%';",
            "SELECT 'after_audit_rows_in_tx', COUNT(*) FROM audit_events WHERE action = 'registry.import' AND target_id = " + sql_str(workbook.name) + ";",
            footer,
            "RESET ROLE;",
        ]
    )
    return "\n\n".join(parts) + "\n"


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--mode", choices=["dry-run", "apply"], required=True)
    parser.add_argument("--target-org-id", default=KNL_ORG_ID)
    parser.add_argument("--target-branch-id", required=True)
    parser.add_argument("--expected-count", type=int, default=444)
    args = parser.parse_args()

    workbook = args.workbook.expanduser()
    if not workbook.exists():
        raise FileNotFoundError(workbook)
    rows = load_rows(workbook)
    if len(rows) != args.expected_count:
        raise ValueError(f"expected {args.expected_count} rows, parsed {len(rows)}")
    sql = build_sql(
        rows,
        workbook=workbook,
        org_id=args.target_org_id,
        branch_id=args.target_branch_id,
        mode=args.mode,
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(sql, encoding="utf-8")
    os.chmod(args.output, 0o600)
    print(
        json.dumps(
            {
                "workbook": str(workbook),
                "output": str(args.output),
                "mode": args.mode,
                "rows": len(rows),
                "sha256": sha256_file(workbook),
                "status_counts": Counter(row["status"] for row in rows),
                "asset_owner_values": len({row["asset_owner"] for row in rows}),
            },
            ensure_ascii=False,
            default=dict,
        )
    )
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:  # noqa: BLE001 - CLI boundary
        print(f"error: {exc}", file=sys.stderr)
        raise SystemExit(1)
