#!/usr/bin/env python3
"""Profile-aware COSS Group workbook import SQL generator.

This script is intentionally local/offline: it reads a folder of Korean HR/
payroll XLSX files, preserves every non-empty workbook row into the restricted
`seed_import` staging schema, writes governed `data_import_runs` /
`data_import_rows` records, and upserts a de-duplicated employee directory.

PII safety:
  * stdout/stderr only contain aggregate counts and file paths.
  * generated SQL contains sensitive source data by design; write it under
    `.omx/context/...` with mode 0600 and never commit it.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import numbers
import os
import re
import sys
import unicodedata
import uuid
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


NAMESPACE = uuid.UUID("8b8e55c8-4d9d-4b0f-9e7f-c65dfcf54b47")

ORG_BY_SLUG = {
    "coss": {"name": "(주)코스"},
    "cnl": {"name": "(주)씨앤엘"},
    "dsl": {"name": "(주)디에스엘"},
    "elso": {"name": "(주)엘소"},
    "knl": {"name": "(주)케이앤엘"},
    "cheongun-hr": {"name": "(주)청운HR"},
    "cheongun-logis": {"name": "(주)청운로지스"},
    "jy-tech": {"name": "제이와이테크"},
}

FOLDER_TO_SLUG = {
    "knl": "knl",
    "케이앤엘": "knl",
    "코스": "coss",
    "코스텍": "coss",
    "씨앤엘오에스": "cnl",
    "씨앤엘": "cnl",
    "디에스엘": "dsl",
    "엘소": "elso",
    "청운hr": "cheongun-hr",
    "청운로지스": "cheongun-logis",
    # The source folder has one generic "청운" workbook but the live 8-org model
    # has no standalone 청운 organization. Keep it under the logistics org so it
    # remains visible in the group while preserving source metadata for later
    # correction if the tenant model expands.
    "청운": "cheongun-logis",
    "제이와이테크": "jy-tech",
}

SHEET_TO_SLUG = {
    "(주)디에스엘": "dsl",
    "디에스엘": "dsl",
    "(주)코스": "coss",
    "코스": "coss",
    "코스텍": "coss",
    "(주)엘소": "elso",
    "엘소": "elso",
    "엘창": "elso",
    "(주)케이앤엘": "knl",
    "케이앤엘": "knl",
    "knl": "knl",
    "(주)씨앤엘": "cnl",
    "씨앤엘": "cnl",
    "씨앤엘오에스": "cnl",
    "(주)청운hr": "cheongun-hr",
    "청운hr": "cheongun-hr",
    "(주)청운로지스": "cheongun-logis",
    "청운로지스": "cheongun-logis",
    "청운": "cheongun-logis",
    "제이와이테크": "jy-tech",
}

NAME_HEADERS = {"성명", "성함", "이름"}
EMPLOYEE_NUMBER_HEADERS = {"사번", "직원번호", "직번"}
ORG_UNIT_HEADERS = {"부서명", "소속", "부서", "근무부서", "도급업체", "업체명"}
JOB_HEADERS = {"업무", "직종", "담당업무", "구분"}
POSITION_HEADERS = {"직책", "직위", "직급"}
WORKSITE_HEADERS = {"근무지", "사업장", "현장", "현장명"}
WORKSITE_ADDRESS_HEADERS = {"근무지(주소)", "근무지주소", "주소", "사업장주소", "현장주소"}
HIRE_DATE_HEADERS = {"입사일", "입사일자", "보험가입일"}
EXIT_DATE_HEADERS = {"퇴사일", "퇴사일자", "보험상실일"}
LEAVE_ACCRUED_HEADERS = {"발생연차"}
LEAVE_USED_HEADERS = {"사용연차"}
LEAVE_REMAINING_HEADERS = {"잔여연차"}
BIRTH_HEADERS = {"생년월일", "생년월일/주민번호"}

RESTRICTED_FRAGMENTS = [
    "주민",
    "급여",
    "시급",
    "통상",
    "수당",
    "국민연금",
    "건강보험",
    "고용보험",
    "산재",
    "소득세",
    "은행",
    "계좌",
    "장애",
    "퇴직금",
    "지급일",
    "급여산정",
    "휴대폰",
    "전화",
    "연락처",
    "개인주소",
    "거주주소",
]

NON_PERSON_VALUES = {
    "성명",
    "성함",
    "이름",
    "합계",
    "총계",
    "소계",
    "계",
    "구분",
    "인원",
    "근무자",
    "성명계",
}


def nfc(value: str) -> str:
    return unicodedata.normalize("NFC", value)


def compact(value: str) -> str:
    return re.sub(r"\s+", "", nfc(value or "")).strip().lower()


def normalize_header(raw: Any, fallback_index: int) -> str:
    text = cell_to_text(raw)
    text = re.sub(r"\s+", "", text)
    if not text:
        return f"__col_{fallback_index}"
    if text in {"성명", "성함", "이름"}:
        return "성명"
    if text == "근무지(주소)":
        return "근무지(주소)"
    return text


def cell_to_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if isinstance(value, (str, bool)):
        return value
    if isinstance(value, numbers.Integral):
        return int(value)
    if isinstance(value, numbers.Real):
        return int(value) if float(value).is_integer() else float(value)
    if isinstance(value, float) and value.is_integer():
        return int(value)
    # openpyxl can surface ArrayFormula/DataTableFormula and other rich cell
    # payloads. Preserve them as source text rather than dropping the cell or
    # failing the whole import.
    return str(value)


def cell_to_text(value: Any) -> str:
    value = cell_to_value(value)
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def sql_str(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + value.replace("'", "''") + "'"


def sql_json(value: Any) -> str:
    return sql_str(json.dumps(value, ensure_ascii=False, separators=(",", ":"))) + "::jsonb"


def sql_uuid(value: uuid.UUID | None) -> str:
    return "NULL" if value is None else sql_str(str(value)) + "::uuid"


def stable_uuid(kind: str, *parts: str) -> uuid.UUID:
    return uuid.uuid5(NAMESPACE, kind + ":" + "|".join(parts))


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def short_hash(value: str, length: int = 24) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()[:length]


def source_label(root: Path, path: Path) -> str:
    return nfc(str(path.relative_to(root)))


def find_header_row(row_values: dict[int, Any]) -> bool:
    normalized = {normalize_header(value, col) for col, value in row_values.items()}
    return bool(normalized & NAME_HEADERS)


def dedupe_headers(row_values: dict[int, Any]) -> dict[int, str]:
    seen: Counter[str] = Counter()
    headers: dict[int, str] = {}
    for col in sorted(row_values):
        header = normalize_header(row_values.get(col), col)
        seen[header] += 1
        if seen[header] > 1:
            header = f"{header}_{seen[header]}"
        headers[col] = header
    return headers


def target_for_header(header: str) -> str | None:
    if header in NAME_HEADERS:
        return "name"
    if header in EMPLOYEE_NUMBER_HEADERS:
        return "employee_number"
    if header in ORG_UNIT_HEADERS:
        return "org_unit"
    if header in JOB_HEADERS:
        return "job"
    if header in POSITION_HEADERS:
        return "position"
    if header in WORKSITE_HEADERS:
        return "worksite_name"
    if header in WORKSITE_ADDRESS_HEADERS:
        return "worksite_address"
    if header in HIRE_DATE_HEADERS:
        return "hire_date"
    if header in EXIT_DATE_HEADERS:
        return "exit_date"
    if header in LEAVE_ACCRUED_HEADERS:
        return "leave_accrued"
    if header in LEAVE_USED_HEADERS:
        return "leave_used"
    if header in LEAVE_REMAINING_HEADERS:
        return "leave_remaining"
    if header in {"회사", "법인", "소속회사"}:
        return "company"
    return None


def column_classification(header: str, target: str | None) -> str:
    if target == "worksite_address" or "주소" in header or "위치" in header:
        return "location"
    if any(fragment in header for fragment in RESTRICTED_FRAGMENTS):
        return "restricted"
    if target is not None:
        return "canonical"
    return "retained"


def mapping_profile(columns: list[dict[str, Any]], source: dict[str, Any]) -> dict[str, Any]:
    return {
        "entity_type": "employee_hr",
        "source": source,
        "target_allowlist": [
            "name",
            "employee_number",
            "org_unit",
            "job",
            "position",
            "worksite_name",
            "worksite_address",
            "hire_date",
            "exit_date",
            "leave_accrued",
            "leave_used",
            "leave_remaining",
            "company",
        ],
        "columns": columns,
        "policy": {
            "unknown_columns": "retain_raw_only",
            "restricted_columns": "retain_raw_mask_preview",
            "blank_name_rows": "preserve_raw_only",
            "server_side_entity_allowlist": ["employee_hr"],
            "person_dedupe_key": "sha256(org_slug|employee_number_or_name_birth_hire_worksite)",
            "stdout_pii": "forbidden",
        },
    }


def infer_slug_from_path(root: Path, path: Path, sheet_name: str) -> str:
    rel_parts = [nfc(part) for part in path.relative_to(root).parts]
    # Expected shape: 2026 / 5월 / subsidiary / workbook.xlsx
    if len(rel_parts) >= 3:
        folder_slug = FOLDER_TO_SLUG.get(compact(rel_parts[2]))
        if folder_slug:
            return folder_slug
    return SHEET_TO_SLUG.get(compact(sheet_name), "coss")


def derive_worksite_from_filename(path: Path) -> str | None:
    stem = nfc(path.stem)
    stem = re.sub(r"\([^)]*\)", "", stem).strip()
    return stem or None


def value_for(raw: dict[str, Any], header_set: set[str]) -> str | None:
    for header in header_set:
        value = raw.get(header)
        text = cell_to_text(value)
        if text:
            return text
    return None


def parse_decimal_text(text: str | None) -> str | None:
    if not text:
        return None
    cleaned = text.replace(",", "").strip()
    if re.fullmatch(r"-?\d+(\.\d+)?", cleaned):
        return cleaned
    return None


def looks_like_person_name(value: str) -> bool:
    text = compact(value)
    if not text or text in NON_PERSON_VALUES:
        return False
    if text.startswith("="):
        return False
    if len(text) > 24:
        return False
    if re.fullmatch(r"[\d,.\-/:]+", text):
        return False
    if any(word in text for word in ["합계", "총계", "소계", "청구", "공제", "지급", "보험료"]):
        return False
    return bool(re.search(r"[가-힣A-Za-z]", text))


@dataclass
class SourceRow:
    org_slug: str
    org_name: str
    rel: str
    file_sha: str
    sheet: str
    row_number: int
    header_row: int | None
    raw: dict[str, Any]
    row_status: str
    source_key: str
    employee_source_key: str | None = None
    canonical: dict[str, Any] | None = None


@dataclass
class WorkbookImport:
    import_id: uuid.UUID
    rel: str
    sha256: str
    sheets: list[dict[str, Any]] = field(default_factory=list)
    staging_rows: list[tuple[str, int, dict[str, Any]]] = field(default_factory=list)


def build_rows(root: Path) -> tuple[list[WorkbookImport], list[SourceRow], dict[str, Any]]:
    files = sorted(path for path in root.rglob("*.xlsx") if not path.name.startswith("~$"))
    workbook_imports: list[WorkbookImport] = []
    import_rows: list[SourceRow] = []
    stats: dict[str, Any] = {
        "files": len(files),
        "sheets": 0,
        "staging_nonempty_rows": 0,
        "data_import_rows": 0,
        "candidate_rows": 0,
        "preserved_rows": 0,
        "canonical_employee_keys": set(),
        "rows_by_org_slug": Counter(),
        "candidate_rows_by_org_slug": Counter(),
        "errors": [],
        "unknown_or_default_coss_rows": 0,
    }

    for path in files:
        rel = source_label(root, path)
        file_sha = sha256_file(path)
        import_id = stable_uuid("seed_import.workbook", rel, file_sha)
        wb_import = WorkbookImport(import_id=import_id, rel=rel, sha256=file_sha)
        try:
            wb = load_workbook(path, read_only=False, data_only=False)
        except Exception as exc:  # pragma: no cover - data dependent
            stats["errors"].append({"rel": rel, "stage": "open", "error_type": type(exc).__name__})
            continue

        try:
            for sheet_index, ws in enumerate(wb.worksheets, start=1):
                stats["sheets"] += 1
                sheet = nfc(ws.title)
                cells_by_row: dict[int, dict[int, Any]] = defaultdict(dict)
                max_col = 0
                for (row_idx, col_idx), cell in ws._cells.items():  # noqa: SLF001 - intentional sparse read
                    value = cell_to_value(cell.value)
                    if value is None or cell_to_text(value) == "":
                        continue
                    cells_by_row[row_idx][col_idx] = value
                    max_col = max(max_col, col_idx)

                header_row: int | None = None
                for row_idx in sorted(idx for idx in cells_by_row if idx <= 12):
                    if find_header_row(cells_by_row[row_idx]):
                        header_row = row_idx
                        break

                headers_by_col = dedupe_headers(cells_by_row.get(header_row, {})) if header_row else {}
                hidden_columns = [
                    col
                    for col, dim in ws.column_dimensions.items()
                    if getattr(dim, "hidden", False)
                ]
                wb_import.sheets.append(
                    {
                        "sheet_name": sheet,
                        "sheet_index": sheet_index,
                        "max_row": max(cells_by_row.keys(), default=0),
                        "max_column": max_col,
                        "header_row": header_row,
                        "headers_json": [headers_by_col[col] for col in sorted(headers_by_col)],
                        "hidden_columns_json": hidden_columns,
                    }
                )

                for row_idx in sorted(cells_by_row):
                    row_values = cells_by_row[row_idx]
                    row_json = {
                        get_column_letter(col): cell_to_value(value)
                        for col, value in sorted(row_values.items())
                    }
                    row_json["__meta"] = {
                        "source_relpath": rel,
                        "sheet": sheet,
                        "row": row_idx,
                    }
                    wb_import.staging_rows.append((sheet, row_idx, row_json))

                    if header_row and row_idx <= header_row:
                        continue

                    org_slug = infer_slug_from_path(root, path, sheet)
                    org_name = ORG_BY_SLUG[org_slug]["name"]
                    if org_slug == "coss" and len(path.relative_to(root).parts) < 4:
                        stats["unknown_or_default_coss_rows"] += 1

                    if headers_by_col:
                        raw = {
                            headers_by_col.get(col, get_column_letter(col)): cell_to_value(value)
                            for col, value in sorted(row_values.items())
                        }
                    else:
                        raw = {
                            get_column_letter(col): cell_to_value(value)
                            for col, value in sorted(row_values.items())
                        }
                    raw["__source"] = {
                        "root": "COSS Group",
                        "relpath": rel,
                        "sha256": file_sha,
                        "sheet": sheet,
                        "row": row_idx,
                        "header_row": header_row,
                        "org_slug": org_slug,
                        "import_profile": "coss_group_profile_v1",
                    }

                    name = value_for(raw, NAME_HEADERS)
                    is_candidate = bool(name and looks_like_person_name(name))
                    source_key = (
                        f"coss-group:file:{file_sha[:16]}:sheet:{short_hash(sheet, 12)}:row:{row_idx}"
                    )
                    row = SourceRow(
                        org_slug=org_slug,
                        org_name=org_name,
                        rel=rel,
                        file_sha=file_sha,
                        sheet=sheet,
                        row_number=row_idx,
                        header_row=header_row,
                        raw=raw,
                        row_status="CANDIDATE" if is_candidate else "PRESERVED",
                        source_key=source_key,
                    )

                    if is_candidate:
                        employee_number = value_for(raw, EMPLOYEE_NUMBER_HEADERS)
                        org_unit = value_for(raw, ORG_UNIT_HEADERS)
                        job = value_for(raw, JOB_HEADERS)
                        position = value_for(raw, POSITION_HEADERS)
                        worksite_name = value_for(raw, WORKSITE_HEADERS) or derive_worksite_from_filename(path)
                        worksite_address = value_for(raw, WORKSITE_ADDRESS_HEADERS)
                        hire_date = value_for(raw, HIRE_DATE_HEADERS)
                        exit_date = value_for(raw, EXIT_DATE_HEADERS)
                        birth = value_for(raw, BIRTH_HEADERS)
                        leave_accrued = parse_decimal_text(value_for(raw, LEAVE_ACCRUED_HEADERS))
                        leave_used = parse_decimal_text(value_for(raw, LEAVE_USED_HEADERS))
                        leave_remaining = parse_decimal_text(value_for(raw, LEAVE_REMAINING_HEADERS))
                        company = value_for(raw, {"회사", "법인", "소속회사"}) or org_name
                        person_basis = "|".join(
                            [
                                org_slug,
                                employee_number or "",
                                compact(name or ""),
                                compact(birth or ""),
                                compact(hire_date or ""),
                                compact(worksite_name or ""),
                            ]
                        )
                        if employee_number:
                            row.employee_source_key = (
                                f"coss-group:employee:{org_slug}:empno:{short_hash(employee_number)}"
                            )
                        else:
                            row.employee_source_key = (
                                f"coss-group:employee:{org_slug}:fp:{short_hash(person_basis)}"
                            )
                        row.canonical = {
                            "company": company,
                            "name": name,
                            "employee_number": employee_number,
                            "org_unit": org_unit,
                            "job": job,
                            "position": position,
                            "worksite_name": worksite_name,
                            "worksite_address": worksite_address,
                            "hire_date": hire_date,
                            "exit_date": exit_date,
                            "employment_status": "EXITED" if exit_date else "ACTIVE",
                            "leave_accrued": leave_accrued,
                            "leave_used": leave_used,
                            "leave_remaining": leave_remaining,
                            "source_key": row.employee_source_key,
                            "source_metadata": {
                                "source_relpath": rel,
                                "source_sha256": file_sha,
                                "source_sheet": sheet,
                                "source_row": row_idx,
                                "row_source_key": source_key,
                                "dedupe": "stable_hash_no_pii_in_key",
                                "import_profile": "coss_group_profile_v1",
                            },
                        }
                        stats["candidate_rows"] += 1
                        stats["candidate_rows_by_org_slug"][org_slug] += 1
                        stats["canonical_employee_keys"].add((org_slug, row.employee_source_key))
                    else:
                        stats["preserved_rows"] += 1

                    stats["rows_by_org_slug"][org_slug] += 1
                    import_rows.append(row)
        except Exception as exc:  # pragma: no cover - data dependent
            stats["errors"].append({"rel": rel, "stage": "parse", "error_type": type(exc).__name__})
        finally:
            try:
                wb.close()
            except Exception:
                pass

        stats["staging_nonempty_rows"] += len(wb_import.staging_rows)
        workbook_imports.append(wb_import)

    stats["data_import_rows"] = len(import_rows)
    stats["canonical_employee_key_count"] = len(stats["canonical_employee_keys"])
    stats["canonical_employee_keys"] = stats["canonical_employee_key_count"]
    stats["rows_by_org_slug"] = dict(stats["rows_by_org_slug"])
    stats["candidate_rows_by_org_slug"] = dict(stats["candidate_rows_by_org_slug"])
    return workbook_imports, import_rows, stats


def build_columns_for_rows(rows: Iterable[SourceRow]) -> list[dict[str, Any]]:
    headers: dict[str, dict[str, Any]] = {}
    for row in rows:
        for header in row.raw:
            if header.startswith("__"):
                continue
            target = target_for_header(header)
            classification = column_classification(header, target)
            headers.setdefault(
                header,
                {
                    "source_header": header,
                    "normalized_header": header,
                    "target": target,
                    "preview_allowed": classification in {"canonical", "retained"},
                    "classification": classification,
                },
            )
    return list(headers.values())


def emit_sql(
    sql_path: Path,
    workbook_imports: list[WorkbookImport],
    rows: list[SourceRow],
    stats: dict[str, Any],
    org_ids: dict[str, str],
) -> None:
    grouped: dict[tuple[str, str, str], list[SourceRow]] = defaultdict(list)
    for row in rows:
        grouped[(row.rel, row.file_sha, row.org_slug)].append(row)

    canonical_by_key: dict[tuple[str, str], SourceRow] = {}
    for row in rows:
        if row.employee_source_key and row.canonical:
            canonical_by_key[(row.org_slug, row.employee_source_key)] = row

    with sql_path.open("w", encoding="utf-8") as out:
        out.write("-- Generated by scripts/import_coss_group_workbooks.py.\n")
        out.write("-- Contains sensitive HR/payroll source data. Do not commit.\n")
        out.write("\\set ON_ERROR_STOP on\n")
        out.write("BEGIN;\n")
        out.write("SET LOCAL statement_timeout = 0;\n")
        out.write("SET LOCAL lock_timeout = '10s';\n")

        for wb in workbook_imports:
            profile = {
                "source_root": "COSS Group",
                "source_relpath": wb.rel,
                "source_sha256": wb.sha256,
                "profile": "coss_group_profile_v1",
                "pii_policy": "staged in database only; aggregate logs only",
            }
            out.write(
                "INSERT INTO seed_import.workbook_imports "
                "(id, source_filename, source_sha256, imported_at, sheet_count, total_rows, total_columns, profile_json) "
                f"VALUES ({sql_uuid(wb.import_id)}, {sql_str(wb.rel)}, {sql_str(wb.sha256)}, now(), "
                f"{len(wb.sheets)}, {len(wb.staging_rows)}, "
                f"{max((sheet['max_column'] for sheet in wb.sheets), default=0)}, {sql_json(profile)}) "
                "ON CONFLICT (id) DO UPDATE SET "
                "imported_at=EXCLUDED.imported_at, sheet_count=EXCLUDED.sheet_count, "
                "total_rows=EXCLUDED.total_rows, total_columns=EXCLUDED.total_columns, "
                "profile_json=EXCLUDED.profile_json;\n"
            )
            for sheet in wb.sheets:
                out.write(
                    "INSERT INTO seed_import.workbook_sheets "
                    "(import_id, sheet_name, sheet_index, max_row, max_column, header_row, headers_json, hidden_columns_json) "
                    f"VALUES ({sql_uuid(wb.import_id)}, {sql_str(sheet['sheet_name'])}, {sheet['sheet_index']}, "
                    f"{sheet['max_row']}, {sheet['max_column']}, "
                    f"{sheet['header_row'] if sheet['header_row'] is not None else 0}, "
                    f"{sql_json(sheet['headers_json'])}, {sql_json(sheet['hidden_columns_json'])}) "
                    "ON CONFLICT (import_id, sheet_name) DO UPDATE SET "
                    "sheet_index=EXCLUDED.sheet_index, max_row=EXCLUDED.max_row, "
                    "max_column=EXCLUDED.max_column, header_row=EXCLUDED.header_row, "
                    "headers_json=EXCLUDED.headers_json, hidden_columns_json=EXCLUDED.hidden_columns_json;\n"
                )
            for sheet_name, row_number, row_json in wb.staging_rows:
                out.write(
                    "INSERT INTO seed_import.workbook_rows (import_id, sheet_name, source_row, row_json) "
                    f"VALUES ({sql_uuid(wb.import_id)}, {sql_str(sheet_name)}, {row_number}, {sql_json(row_json)}) "
                    "ON CONFLICT (import_id, sheet_name, source_row) DO UPDATE SET row_json=EXCLUDED.row_json;\n"
                )

        for (rel, file_sha, org_slug), run_rows in grouped.items():
            org_id = org_ids[org_slug]
            run_id = stable_uuid("data_import_run", rel, file_sha, org_slug)
            candidate_rows = [row for row in run_rows if row.row_status == "CANDIDATE"]
            preserved_rows = [row for row in run_rows if row.row_status != "CANDIDATE"]
            columns = build_columns_for_rows(run_rows)
            dry_run = {
                "run_id": str(run_id),
                "input_rows": len(run_rows),
                "candidate_rows": len(candidate_rows),
                "preserved_rows": len(preserved_rows),
                "insert_candidates": len({row.employee_source_key for row in candidate_rows if row.employee_source_key}),
                "update_candidates": 0,
                "mode": "profile_aware_direct_apply",
            }
            apply_summary = {
                "input_rows": len(run_rows),
                "candidate_rows": len(candidate_rows),
                "preserved_rows": len(preserved_rows),
                "canonical_employee_keys": len({row.employee_source_key for row in candidate_rows if row.employee_source_key}),
                "mode": "profile_aware_direct_apply",
            }
            out.write(
                "INSERT INTO data_import_runs "
                "(id, org_id, entity_type, status, source_filename, source_format, source_sha256, "
                "mapping_profile, dry_run_summary, apply_summary, input_rows, candidate_rows, preserved_rows, "
                "created_by, applied_by, created_at, updated_at, applied_at) VALUES ("
                f"{sql_uuid(run_id)}, {sql_uuid(uuid.UUID(org_id))}, 'employee_hr', 'APPLIED', "
                f"{sql_str(rel)}, 'xlsx', {sql_str(file_sha)}, "
                f"{sql_json(mapping_profile(columns, {'relpath': rel, 'org_slug': org_slug}))}, "
                f"{sql_json(dry_run)}, {sql_json(apply_summary)}, "
                f"{len(run_rows)}, {len(candidate_rows)}, {len(preserved_rows)}, "
                "NULL, NULL, now(), now(), now()) "
                "ON CONFLICT (id) DO UPDATE SET "
                "status='APPLIED', mapping_profile=EXCLUDED.mapping_profile, "
                "dry_run_summary=EXCLUDED.dry_run_summary, apply_summary=EXCLUDED.apply_summary, "
                "input_rows=EXCLUDED.input_rows, candidate_rows=EXCLUDED.candidate_rows, "
                "preserved_rows=EXCLUDED.preserved_rows, updated_at=now(), applied_at=now();\n"
            )
            for row in run_rows:
                if row.canonical and row.employee_source_key:
                    canonical_row = {
                        "company": row.canonical["company"],
                        "name": row.canonical["name"],
                        "source_filename": row.rel,
                        "source_sheet": row.sheet,
                        "source_row": row.row_number,
                        "source_key": row.employee_source_key,
                        "source_metadata": row.canonical["source_metadata"],
                        "canonical": {
                            key: row.canonical.get(key)
                            for key in [
                                "employee_number",
                                "org_unit",
                                "job",
                                "position",
                                "worksite_name",
                                "worksite_address",
                                "hire_date",
                                "exit_date",
                                "employment_status",
                                "leave_accrued",
                                "leave_used",
                                "leave_remaining",
                            ]
                        },
                    }
                    validation = {"status": "ok", "errors": [], "warnings": []}
                else:
                    canonical_row = {
                        "source_filename": row.rel,
                        "source_sheet": row.sheet,
                        "source_row": row.row_number,
                        "source_key": row.source_key,
                        "raw_only_reason": "missing_or_non_person_name",
                    }
                    validation = {
                        "status": "preserved",
                        "errors": [],
                        "warnings": ["missing_or_non_person_name_preserved_raw_only"],
                    }
                out.write(
                    "INSERT INTO data_import_rows "
                    "(org_id, run_id, source_sheet, source_row, source_key, row_status, raw_row, canonical_row, validation) VALUES ("
                    f"{sql_uuid(uuid.UUID(org_id))}, {sql_uuid(run_id)}, {sql_str(row.sheet)}, {row.row_number}, "
                    f"{sql_str(row.source_key)}, {sql_str(row.row_status)}, {sql_json(row.raw)}, "
                    f"{sql_json(canonical_row)}, {sql_json(validation)}) "
                    "ON CONFLICT (run_id, source_key) DO NOTHING;\n"
                )

        for (org_slug, employee_key), row in sorted(canonical_by_key.items()):
            assert row.canonical is not None
            org_id = org_ids[org_slug]
            c = row.canonical
            out.write(
                "INSERT INTO employees "
                "(org_id, company, name, source_filename, source_sheet, source_row, source_key, raw_row, "
                "source_metadata, employee_number, org_unit, job, position, worksite_name, worksite_address, "
                "hire_date, exit_date, employment_status, leave_accrued, leave_used, leave_remaining) VALUES ("
                f"{sql_uuid(uuid.UUID(org_id))}, {sql_str(c['company'])}, {sql_str(c['name'])}, "
                f"{sql_str(row.rel)}, {sql_str(row.sheet)}, {row.row_number}, {sql_str(employee_key)}, "
                f"{sql_json(row.raw)}, {sql_json(c['source_metadata'])}, "
                f"{sql_str(c.get('employee_number'))}, {sql_str(c.get('org_unit'))}, {sql_str(c.get('job'))}, "
                f"{sql_str(c.get('position'))}, {sql_str(c.get('worksite_name'))}, {sql_str(c.get('worksite_address'))}, "
                f"{sql_str(c.get('hire_date'))}, {sql_str(c.get('exit_date'))}, {sql_str(c.get('employment_status') or 'ACTIVE')}, "
                f"NULLIF({sql_str(c.get('leave_accrued'))}, '')::NUMERIC, "
                f"NULLIF({sql_str(c.get('leave_used'))}, '')::NUMERIC, "
                f"NULLIF({sql_str(c.get('leave_remaining'))}, '')::NUMERIC) "
                "ON CONFLICT (org_id, source_key) DO UPDATE SET "
                "company=EXCLUDED.company, name=EXCLUDED.name, source_filename=EXCLUDED.source_filename, "
                "source_sheet=EXCLUDED.source_sheet, source_row=EXCLUDED.source_row, raw_row=EXCLUDED.raw_row, "
                "source_metadata=EXCLUDED.source_metadata, employee_number=EXCLUDED.employee_number, "
                "org_unit=EXCLUDED.org_unit, job=EXCLUDED.job, position=EXCLUDED.position, "
                "worksite_name=EXCLUDED.worksite_name, worksite_address=EXCLUDED.worksite_address, "
                "hire_date=EXCLUDED.hire_date, exit_date=EXCLUDED.exit_date, "
                "employment_status=EXCLUDED.employment_status, leave_accrued=EXCLUDED.leave_accrued, "
                "leave_used=EXCLUDED.leave_used, leave_remaining=EXCLUDED.leave_remaining, updated_at=now();\n"
            )

        audit_id = stable_uuid("audit.coss_group_import", stats.get("created_at_utc", ""), str(stats["files"]))
        after = {
            "source": "COSS Group",
            "files": stats["files"],
            "sheets": stats["sheets"],
            "staging_nonempty_rows": stats["staging_nonempty_rows"],
            "data_import_rows": stats["data_import_rows"],
            "candidate_rows": stats["candidate_rows"],
            "preserved_rows": stats["preserved_rows"],
            "canonical_employee_keys": stats["canonical_employee_key_count"],
            "rows_by_org_slug": stats["rows_by_org_slug"],
            "pii_values_returned": False,
        }
        out.write(
            "INSERT INTO audit_events "
            "(id, actor, action, target_type, target_id, branch_id, before_snap, after_snap, trace_id, span_id, occurred_at, created_at, org_id) "
            f"VALUES ({sql_uuid(audit_id)}, NULL, 'data_import.profile_aware_apply', 'data_import_batch', "
            f"{sql_str('COSS Group 2026-05')}, NULL, NULL, {sql_json(after)}, "
            f"{sql_str(short_hash(str(audit_id), 32))}, {sql_str(short_hash(str(audit_id) + ':span', 16))}, now(), now(), NULL) "
            "ON CONFLICT (id) DO NOTHING;\n"
        )

        out.write("COMMIT;\n")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", default="/Users/jasonlee/Desktop/COSS Group")
    parser.add_argument("--summary-out", required=True)
    parser.add_argument("--sql-out", required=True)
    parser.add_argument(
        "--org-id",
        action="append",
        default=[],
        help="slug=uuid mapping from live organizations table; repeatable",
    )
    args = parser.parse_args()

    source = Path(args.source).expanduser().resolve()
    if not source.exists():
        raise SystemExit(f"source not found: {source}")

    org_ids: dict[str, str] = {}
    for item in args.org_id:
        slug, _, value = item.partition("=")
        if not slug or not value:
            raise SystemExit(f"invalid --org-id mapping: {item}")
        uuid.UUID(value)
        org_ids[slug] = value
    missing = sorted(set(ORG_BY_SLUG) - set(org_ids))
    if missing:
        raise SystemExit(f"missing --org-id mappings: {', '.join(missing)}")

    workbook_imports, rows, stats = build_rows(source)
    stats["created_at_utc"] = dt.datetime.now(dt.timezone.utc).isoformat()
    stats["source_root"] = str(source)
    stats["pii_policy"] = "Summary omits row values and names; SQL artifact contains sensitive source data."

    summary_path = Path(args.summary_out)
    sql_path = Path(args.sql_out)
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    sql_path.parent.mkdir(parents=True, exist_ok=True)

    emit_sql(sql_path, workbook_imports, rows, stats, org_ids)
    os.chmod(sql_path, 0o600)
    summary_path.write_text(json.dumps(stats, ensure_ascii=False, indent=2), encoding="utf-8")

    print(
        json.dumps(
            {
                "files": stats["files"],
                "sheets": stats["sheets"],
                "staging_nonempty_rows": stats["staging_nonempty_rows"],
                "data_import_rows": stats["data_import_rows"],
                "candidate_rows": stats["candidate_rows"],
                "preserved_rows": stats["preserved_rows"],
                "canonical_employee_keys": stats["canonical_employee_key_count"],
                "rows_by_org_slug": stats["rows_by_org_slug"],
                "errors_count": len(stats["errors"]),
                "summary_path": str(summary_path),
                "sql_path": str(sql_path),
            },
            ensure_ascii=False,
        )
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
